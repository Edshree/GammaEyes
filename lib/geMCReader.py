# -*- coding: utf-8 -*-

# GammaEyes
#
# Created at: 2021.07.19
#
# A class for gamma spectrum
import numpy as np
import openpyxl as oe
from PyQt5.QtWidgets import QMessageBox


class geMCReader:
    def __init__(self):
        self.stopSym = '      total'

    def chrExcel(self,x):
        if x >= 65 and x <= 90:
            vName = chr(x)
        elif x > 90 and x <= 116:
            vName = 'A' + chr(x - 26)
        else:
            QMessageBox.warning(self,
                                "Error!",
                                "0000")
        return vName

    def open(self, fn_list,prcb):
        self.wb = oe.Workbook()

        for i in range(len(fn_list)):
            # print(fn_list[i])

            proF = (i + 1.0) / len(fn_list) * 100

            wsName = fn_list[i].split('/')[-1]
            ws = self.wb.create_sheet(wsName)

            kline = 0

            fcChar = []  # 记录f卡类别的list
            for line in open(fn_list[i]):
                kline += 1
                '''=========================
                               1
                   =========================        
                '''
                # fCardStr = line.split(' ')

                line14 = line[14:]
                if len(line) > 14:
                    if (line[13] == 'F' or line[13] == 'f') and line.find(':') >= 0:  # 判断是否为f卡
                        fChara = line[14:line.find(':')]  # f卡特征值
                        fClass = fChara[-1]  # f卡类别

                        fNum = line[line.find(':') + 1:].split(' ')
                        for j in range(1, len(fNum)):
                            if fNum[j] and fNum[j] != '\n':
                                fcChar.append(fClass + '-' + fNum[0] + '-' + fNum[j])
            klineAll = kline
            '''    第二次循环
            '''

            for j in range(len(fcChar)):

                fc = fcChar[j].split('-')
                if fc[0] == '1' or fc[0] == '2':
                    iden0 = ' surface  '
                elif fc[0] == '4' or fc[0] == '8':
                    iden0 = ' cell  '
                else:
                    QMessageBox.warning(self,
                                        "Error!",
                                        fc[0] + "card is not support")

                iden = iden0 + fc[2]
                # print(iden)

                letter1 = self.chrExcel(j * 4 + 65)
                letter2 = self.chrExcel(j * 4 + 66)
                letter3 = self.chrExcel(j * 4 + 67)

                wsvName1 = letter1 + '1'
                wsvName2 = letter2 + '1'
                wsvName3 = letter3 + '1'
                ws[wsvName1] = 'card:' + fc[0]
                ws[wsvName2] = 'par:' + fc[1]
                ws[wsvName3] = 'vol:' + fc[2]

                prcb.setValue(int(proF))

                kline = 0
                dataLine = klineAll + 1

                wsRowInt = 1
                stopK = 0

                for line in open(fn_list[i]):
                    kline += 1
                    # print(kline,dataLine,stopK)

                    if line[:len(iden)] == iden:
                        dataLine = kline + 2
                        # print(dataLine)
                    if line[:11] == self.stopSym:
                        dataLine = klineAll + 1
                        stopK += 1
                        # print(j,'stop')

                    if kline >= dataLine and stopK <= j:
                        wsRowInt += 1

                        wsRow1 = letter1 + str(wsRowInt)
                        wsRow2 = letter2 + str(wsRowInt)
                        wsRow3 = letter3 + str(wsRowInt)

                        data = line.split(' ')
                        ener = data[4]
                        distr = data[7]
                        cred = data[8][:-2]

                        ws[wsRow1] = ener
                        ws[wsRow2] = distr
                        ws[wsRow3] = cred

                        # print(ener, distr, cred)



    def save(self, excel_path):
        print(excel_path)
        self.wb.save(excel_path)







